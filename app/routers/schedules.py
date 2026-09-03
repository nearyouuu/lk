from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import select, or_, distinct
from sqlalchemy import text
from sqlalchemy.exc import OperationalError, ProgrammingError
from datetime import datetime
from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.deps import get_db, get_current_user, require_permission, is_admin
from app.models.schedule import (
    Lesson, Group, Teacher, Subject, Room, LessonTime, lesson_teachers, teacher_subjects,
)
from app.models.grade import Student as StudentModel, Grade
from app.models.user import User
from app.schemas.schedule import LessonCreate, LessonOut, SubjectType
from app.models.role import Role, user_roles
from app.schemas.schedule import LessonTopicUpdate, LessonUpdate
from app.services.subject_service import resolve_subject
from app.services.subject_teacher_service import (
    ensure_teacher_linked_to_subject,
    subject_teacher_payload,
)

router = APIRouter(prefix="/schedules", tags=["schedules"])

SCHEDULE_EXPORT_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SUBJECT_TYPE_LABELS = {
    "lecture": "Лекция",
    "practice": "Практика",
    "lab": "Лабораторная",
}

def get_or_create(db, model, where: dict, defaults: dict = {}):
    inst = db.scalar(select(model).filter_by(**where))
    if inst:
        return inst
    inst = model(**where, **defaults)
    db.add(inst); db.flush()
    return inst

def user_has_role(db: Session, user_id: int, role_name: str) -> bool:
    """Проверка роли пользователя."""
    q = (
        select(Role)
        .join(user_roles, Role.id == user_roles.c.role_id)
        .where(user_roles.c.user_id == user_id, Role.name == role_name)
    )
    return db.scalar(q) is not None

@router.post("/lessons", dependencies=[Depends(require_permission("schedules:create"))])
def create_lesson(payload: LessonCreate, db: Session = Depends(get_db), me=Depends(get_current_user)):
    group = get_or_create(db, Group, {"code": payload.group_code}, {"title": payload.group_code})

    if payload.subject_code or payload.subject_id is not None:
        subject = resolve_subject(db, payload.subject_code, payload.subject_id)
    else:
        subject = get_or_create(db, Subject, {"title": payload.subject_title})

    if payload.teacher_id is not None:
        teacher = db.get(Teacher, payload.teacher_id)
        if not teacher:
            raise HTTPException(status_code=404, detail="Teacher not found")
    else:
        teacher = get_or_create(
            db, Teacher,
            {"full_name": payload.teacher_full_name},
            {"email": payload.teacher_email, "phone": payload.teacher_phone, "subject": payload.teacher_subject}
        )

    ensure_teacher_linked_to_subject(db, teacher.id, subject.id)

    room = get_or_create(db, Room, {"code": payload.room_code}, {"title": payload.room_code})

    if payload.lesson_number is not None:
        lesson_time = db.scalar(
            select(LessonTime).where(LessonTime.lesson_number == payload.lesson_number)
        )
    else:
        lesson_time = db.scalar(
            select(LessonTime)
            .where(
                LessonTime.start_time <= payload.starts_at.time(),
                LessonTime.end_time >= payload.starts_at.time(),
            )
            .order_by(LessonTime.lesson_number)
        )
    if lesson_time is None:
        raise HTTPException(status_code=422, detail="LessonTime not found")

    lesson = Lesson(
        group_id=group.id, subject_id=subject.id, teacher_id=teacher.id, room_id=room.id,
        lesson_number=lesson_time.lesson_number,
        starts_at=payload.starts_at, ends_at=payload.ends_at,
        subject_type=payload.subject_type, topic=payload.topic, notes=payload.notes, created_by=me.id
    )
    db.add(lesson); db.commit()
    return {"id": lesson.id, "subject_code": subject.code}

@router.post(
    "/lessons/{lesson_id}",
    dependencies=[Depends(require_permission("schedules:update"))]
)
def update_lesson(
    lesson_id: int,
    payload: LessonUpdate,
    db: Session = Depends(get_db),
    me=Depends(get_current_user),
):
    lesson = db.get(Lesson, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")

    if payload.group_code:
        group = get_or_create(db, Group, {"code": payload.group_code}, {"title": payload.group_code})
        lesson.group_id = group.id

    if payload.subject_code or payload.subject_id is not None:
        subject = resolve_subject(db, payload.subject_code, payload.subject_id)
        lesson.subject_id = subject.id
    elif payload.subject_title:
        subject = get_or_create(db, Subject, {"title": payload.subject_title})
        lesson.subject_id = subject.id

    if payload.teacher_id is not None:
        teacher = db.get(Teacher, payload.teacher_id)
        if not teacher:
            raise HTTPException(status_code=404, detail="Teacher not found")
    elif payload.teacher_full_name:
        teacher = get_or_create(
            db, Teacher,
            {"full_name": payload.teacher_full_name},
            {"email": payload.teacher_email, "phone": payload.teacher_phone, "subject": payload.teacher_subject}
        )
    else:
        teacher = None

    if teacher:
        subject = db.get(Subject, lesson.subject_id)
        if subject:
            ensure_teacher_linked_to_subject(db, teacher.id, subject.id)
        lesson.teacher_id = teacher.id
    elif lesson.teacher_id is not None and lesson.subject_id is not None:
        ensure_teacher_linked_to_subject(db, lesson.teacher_id, lesson.subject_id)

    if payload.room_code:
        room = get_or_create(db, Room, {"code": payload.room_code}, {"title": payload.room_code})
        lesson.room_id = room.id

    if payload.starts_at is not None:
        lesson.starts_at = payload.starts_at
    if payload.ends_at is not None:
        lesson.ends_at = payload.ends_at
    if payload.subject_type is not None:
        lesson.subject_type = payload.subject_type
    if "topic" in payload.model_fields_set:
        lesson.topic = payload.topic
    if payload.notes is not None:
        lesson.notes = payload.notes
    if payload.lesson_number is not None:
        lesson.lesson_number = payload.lesson_number

    db.commit()
    db.refresh(lesson)
    return {
        "id": lesson.id,
        "subject_code": lesson.subject.code if lesson.subject else None,
        "updated": True,
    }

@router.get("/teachers/{teacher_id}/teaching", dependencies=[Depends(require_permission("schedules:read"))])
def teacher_teaching_overview(
    teacher_id: int,
    db: Session = Depends(get_db),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")

    groups_stmt = (
        select(distinct(Group.id), Group.identifier, Group.code, Group.title)
        .join(Lesson, Lesson.group_id == Group.id)
        .join(teacher_subjects, teacher_subjects.c.subject_id == Lesson.subject_id)
        .where(teacher_subjects.c.teacher_id == teacher_id)
    )
    if date_from:
        groups_stmt = groups_stmt.where(Lesson.starts_at >= date_from)
    if date_to:
        groups_stmt = groups_stmt.where(Lesson.starts_at < date_to)
    groups_rows = db.execute(groups_stmt).all()
    groups = [
        {"id": gid, "identifier": identifier, "code": gcode, "title": gtitle}
        for gid, identifier, gcode, gtitle in groups_rows
    ]
    group_ids = [g["id"] for g in groups] or [-1]

    subj_stmt = (
        select(distinct(Group.id), Subject.id, Subject.title, Subject.code, Subject.grade_type)
        .join(Lesson, Lesson.group_id == Group.id)
        .join(Subject, Subject.id == Lesson.subject_id)
        .join(teacher_subjects, teacher_subjects.c.subject_id == Subject.id)
        .where(
            teacher_subjects.c.teacher_id == teacher_id,
            Lesson.group_id.in_(group_ids),
        )
    )
    if date_from:
        subj_stmt = subj_stmt.where(Lesson.starts_at >= date_from)
    if date_to:
        subj_stmt = subj_stmt.where(Lesson.starts_at < date_to)
    subj_rows = db.execute(subj_stmt).all()
    subjects_by_group = {}
    for gid, sid, stitle, scode, grade_type in subj_rows:
        subjects_by_group.setdefault(gid, []).append({"id": sid, "title": stitle, "code": scode, "subject_code": scode, "grade_type": grade_type})

    lesson_stmt = (
        select(Lesson)
        .where(Lesson.teacher_id == teacher_id, Lesson.group_id.in_(group_ids))
        .order_by(Lesson.starts_at)
    )
    if date_from:
        lesson_stmt = lesson_stmt.where(Lesson.starts_at >= date_from)
    if date_to:
        lesson_stmt = lesson_stmt.where(Lesson.starts_at < date_to)
    lessons = db.scalars(lesson_stmt).all()

    final_grades_stmt = (
        select(
            Grade.student_id,
            Grade.subject_id,
            Grade.value,
            StudentModel.group_id,
            User.full_name,
            Subject.title.label("subject_title"),
            Subject.code.label("subject_code"),
            Group.code.label("group_code"),
        )
        .join(StudentModel, StudentModel.id == Grade.student_id)
        .join(User, User.id == StudentModel.user_id)
        .join(Subject, Subject.id == Grade.subject_id)
        .join(Group, Group.id == StudentModel.group_id)
        .where(
            Grade.teacher_id == teacher_id,
            Grade.grade_type.in_(["итог", "final", "exam", "зачет"]),
            StudentModel.group_id.in_(group_ids)
        )
    )
    final_grades_rows = db.execute(final_grades_stmt).all()

    final_grades_by_group = {}
    for row in final_grades_rows:
        gid = row.group_id
        sid = row.subject_id
        final_grades_by_group.setdefault(gid, {}).setdefault(sid, []).append({
            "student_id": row.student_id,
            "student_name": row.full_name,
            "value": row.value,
            "subject_title": row.subject_title,
            "subject_code": row.subject_code,
            "group_code": row.group_code,
        })

    return {
        "teacher_id": teacher.id,
        "teacher_full_name": teacher.full_name,
        "groups": groups,
        "subjectsByGroup": [
            {
                "group": next((g for g in groups if g["id"] == gid), {"id": gid}),
                "subjects": subjects_by_group.get(gid, [])
            }
            for gid in group_ids if gid != -1
        ],
        "lessons": [
            {
                "id": l.id,
                "group_id": l.group_id,
                "group_code": l.group.code if l.group else None,
                "subject_id": l.subject_id,
                "subject_code": l.subject.code if l.subject else None,
                "subject_title": l.subject.title if l.subject else None,
                "room": l.room.code if l.room else None,
                "starts_at": l.starts_at,
                "ends_at": l.ends_at,
                "subject_type": l.subject_type,
                "topic": l.topic,
                "notes":l.notes,
            } for l in lessons
        ],
        "final_grades": [
            {
                "group_id": gid,
                "group_code": next((g["code"] for g in groups if g["id"] == gid), None),
                "subjects": [
                    {
                        "subject_id": sid,
                        "subject_code": next(
                            (s["code"] for s in subjects_by_group.get(gid, []) if s["id"] == sid),
                            None,
                        ),
                        "subject_title": next(
                            (s["title"] for s in subjects_by_group.get(gid, []) if s["id"] == sid),
                            None
                        ),
                        "students": final_grades_by_group[gid][sid],
                    }
                    for sid in final_grades_by_group[gid]
                ],
            }
            for gid in final_grades_by_group
        ],
    }

@router.get("/lessons", response_model=list[LessonOut], dependencies=[Depends(require_permission("schedules:read"))])
def list_lessons(
    db: Session = Depends(get_db),
    group_code: str | None = None,
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    teacher_full_name: str | None = Query(None, description="ФИО преподавателя (подстрочный поиск)"),
    subject_code: str | None = Query(None, description="Код дисциплины"),
    subject_title: str | None = Query(None, description="Название предмета (подстрочный поиск)"),
    room_code: str | None = Query(None, description="Код аудитории"),
    subject_type: SubjectType | None = Query(None, description="lecture/practice/lab"),
):
    try:
        q = (
            select(Lesson)
            .join(Group, Lesson.group_id == Group.id)
            .join(Subject, Lesson.subject_id == Subject.id)
            .outerjoin(Teacher, Lesson.teacher_id == Teacher.id)
            .outerjoin(Room, Lesson.room_id == Room.id)
        )
        if group_code:
            q = q.where(Group.code == group_code)
        if date_from:
            q = q.where(Lesson.starts_at >= date_from)
        if date_to:
            q = q.where(Lesson.starts_at < date_to)
        if teacher_full_name:
            q = q.where(Teacher.full_name.ilike(f"%{teacher_full_name}%"))
        if subject_title:
            q = q.where(Subject.title.ilike(f"%{subject_title}%"))
        if subject_code:
            q = q.where(Subject.code == subject_code.strip())
        if room_code:
            q = q.where(Room.code == room_code)
        if subject_type:
            q = q.where(Lesson.subject_type == subject_type)

        rows = db.scalars(q).all()

        return [
            LessonOut(
                id=l.id,
                group=l.group.code if l.group else None,
                subject=l.subject.title if l.subject else None,
                subject_code=l.subject.code if l.subject else None,
                teacher=l.teacher.full_name if l.teacher else None,
                room=l.room.code if l.room else None,
                starts_at=l.starts_at,
                ends_at=l.ends_at,
                subject_type=l.subject_type,
                topic=l.topic,
                notes=l.notes,
                lesson_number=l.lesson_number,
            )
            for l in rows
        ]
    except (OperationalError, ProgrammingError):
        db.rollback()

        sql = """
            SELECT
                l.id,
                g.code AS group_code,
                s.title AS subject_title,
                s.code AS subject_code,
                t.full_name AS teacher_name,
                r.code AS room_code,
                l.starts_at,
                l.ends_at,
                l.subject_type,
                l.topic,
                l.notes
            FROM lessons l
            JOIN groups g ON l.group_id = g.id
            JOIN subjects s ON l.subject_id = s.id
            LEFT JOIN teachers t ON l.teacher_id = t.id
            LEFT JOIN rooms r ON l.room_id = r.id
        """

        clauses: list[str] = []
        params: dict[str, object] = {}
        if group_code:
            clauses.append("g.code = :group_code")
            params["group_code"] = group_code
        if date_from:
            clauses.append("l.starts_at >= :date_from")
            params["date_from"] = date_from
        if date_to:
            clauses.append("l.starts_at < :date_to")
            params["date_to"] = date_to
        if teacher_full_name:
            clauses.append("t.full_name ILIKE :teacher_full_name")
            params["teacher_full_name"] = f"%{teacher_full_name}%"
        if subject_title:
            clauses.append("s.title ILIKE :subject_title")
            params["subject_title"] = f"%{subject_title}%"
        if subject_code:
            clauses.append("s.code = :subject_code")
            params["subject_code"] = subject_code.strip()
        if room_code:
            clauses.append("r.code = :room_code")
            params["room_code"] = room_code
        if subject_type:
            clauses.append("l.subject_type = :subject_type")
            params["subject_type"] = subject_type

        if clauses:
            sql += " WHERE " + " AND ".join(clauses)
        sql += " ORDER BY l.starts_at"

        rows = db.execute(text(sql), params).mappings().all()
        return [
            LessonOut(
                id=row["id"],
                group=row["group_code"],
                subject=row["subject_title"],
                subject_code=row["subject_code"],
                teacher=row["teacher_name"],
                room=row["room_code"],
                starts_at=row["starts_at"],
                ends_at=row["ends_at"],
                subject_type=row["subject_type"],
                topic=row["topic"],
                notes=row["notes"],
                lesson_number=None,
            )
            for row in rows
        ]


@router.get(
    "/export",
    dependencies=[Depends(require_permission("schedules:read"))],
    summary="Скачать расписание для печати",
)
def export_schedule(
    db: Session = Depends(get_db),
    group_code: str | None = None,
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    teacher_full_name: str | None = Query(None, description="ФИО преподавателя (подстрочный поиск)"),
    subject_code: str | None = Query(None, description="Код дисциплины"),
    subject_title: str | None = Query(None, description="Название предмета (подстрочный поиск)"),
    room_code: str | None = Query(None, description="Код аудитории"),
    subject_type: SubjectType | None = Query(None, description="lecture/practice/lab"),
):
    """Выгрузить отфильтрованное расписание в подготовленный для печати Excel-файл."""
    lessons = list_lessons(
        db=db,
        group_code=group_code,
        date_from=date_from,
        date_to=date_to,
        teacher_full_name=teacher_full_name,
        subject_code=subject_code,
        subject_title=subject_title,
        room_code=room_code,
        subject_type=subject_type,
    )
    lessons = sorted(lessons, key=lambda lesson: (lesson.starts_at, lesson.lesson_number or 0))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Расписание"
    sheet.sheet_view.showGridLines = False

    columns = [
        ("Дата", 13),
        ("День недели", 15),
        ("Пара", 8),
        ("Время", 14),
        ("Группа", 13),
        ("Дисциплина", 32),
        ("Тип", 17),
        ("Преподаватель", 28),
        ("Аудитория", 13),
        ("Тема урока", 30),
        ("Примечание", 25),
    ]
    last_column = get_column_letter(len(columns))

    sheet.merge_cells(f"A1:{last_column}1")
    title = sheet["A1"]
    title.value = "Расписание занятий"
    title.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    filter_parts = []
    if group_code:
        filter_parts.append(f"Группа: {group_code}")
    if date_from:
        filter_parts.append(f"с {date_from.strftime('%d.%m.%Y')}")
    if date_to:
        filter_parts.append(f"до {date_to.strftime('%d.%m.%Y')}")
    if teacher_full_name:
        filter_parts.append(f"Преподаватель: {teacher_full_name}")
    if subject_code:
        filter_parts.append(f"Код дисциплины: {subject_code}")
    if subject_title:
        filter_parts.append(f"Дисциплина: {subject_title}")
    if room_code:
        filter_parts.append(f"Аудитория: {room_code}")
    if subject_type:
        filter_parts.append(f"Тип: {SUBJECT_TYPE_LABELS.get(subject_type, subject_type)}")

    sheet.merge_cells(f"A2:{last_column}2")
    subtitle = sheet["A2"]
    subtitle.value = " • ".join(filter_parts) if filter_parts else "Все занятия"
    subtitle.font = Font(name="Arial", size=10, italic=True, color="44546A")
    subtitle.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 24

    header_row = 4
    thin_gray = Side(style="thin", color="B4C6E7")
    for column_number, (label, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=column_number, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="1F4E78"))
        sheet.column_dimensions[get_column_letter(column_number)].width = width
    sheet.row_dimensions[header_row].height = 30

    weekday_names = (
        "Понедельник",
        "Вторник",
        "Среда",
        "Четверг",
        "Пятница",
        "Суббота",
        "Воскресенье",
    )
    first_data_row = header_row + 1
    for row_number, lesson in enumerate(lessons, start=first_data_row):
        values = [
            lesson.starts_at.date(),
            weekday_names[lesson.starts_at.weekday()],
            lesson.lesson_number,
            f"{lesson.starts_at.strftime('%H:%M')}–{lesson.ends_at.strftime('%H:%M')}",
            lesson.group,
            lesson.subject,
            SUBJECT_TYPE_LABELS.get(lesson.subject_type, lesson.subject_type or ""),
            lesson.teacher or "",
            lesson.room or "",
            lesson.topic or "",
            lesson.notes or "",
        ]
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column_number, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                horizontal="left" if column_number in {6, 8, 10, 11} else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=thin_gray)
            if row_number % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF2F8")
        sheet.cell(row=row_number, column=1).number_format = "dd.mm.yyyy"
        sheet.row_dimensions[row_number].height = 30

    if not lessons:
        sheet.merge_cells(start_row=first_data_row, start_column=1, end_row=first_data_row, end_column=len(columns))
        empty_cell = sheet.cell(row=first_data_row, column=1, value="По выбранным фильтрам занятий нет")
        empty_cell.font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[first_data_row].height = 32

    last_row = max(first_data_row, sheet.max_row)
    sheet.freeze_panes = f"A{first_data_row}"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:{last_column}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "Страница &P из &N"
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2

    stream = BytesIO()
    workbook.save(stream)
    filename = f"raspisanie_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    encoded_filename = quote(filename)

    return Response(
        content=stream.getvalue(),
        media_type=SCHEDULE_EXPORT_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="schedule.xlsx"; filename*=UTF-8\'\'{encoded_filename}'
            )
        },
    )


@router.post(
    "/lessons/{lesson_id}/topic",
    summary="Изменить тему урока",
)
def update_lesson_topic(
    lesson_id: int,
    payload: LessonTopicUpdate,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    lesson = db.get(Lesson, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")

    privileged = is_admin(me, db) or user_has_role(db, me.id, "director")
    if not privileged:
        if not user_has_role(db, me.id, "teacher"):
            raise HTTPException(status_code=403, detail="Only teachers can update a lesson topic")

        teacher = db.scalar(select(Teacher).where(Teacher.user_id == me.id))
        if not teacher:
            raise HTTPException(status_code=403, detail="Teacher profile not found")

        assigned_as_additional = db.scalar(
            select(lesson_teachers.c.lesson_id).where(
                lesson_teachers.c.lesson_id == lesson.id,
                lesson_teachers.c.teacher_id == teacher.id,
            )
        )
        if lesson.teacher_id != teacher.id and assigned_as_additional is None:
            raise HTTPException(status_code=403, detail="Teacher is not assigned to this lesson")

    lesson.topic = payload.topic
    db.commit()
    db.refresh(lesson)
    return {"id": lesson.id, "topic": lesson.topic, "updated": True}

@router.get("/lookup/groups")
def lookup_groups(db: Session = Depends(get_db), q: str | None = Query(None)):
    stmt = select(Group)
    if q:
        stmt = stmt.where(or_(
            Group.identifier.ilike(f"%{q}%"),
            Group.code.ilike(f"%{q}%"),
            Group.title.ilike(f"%{q}%"),
        ))
    rows = db.scalars(stmt).all()
    return [
        {"id": g.id, "identifier": g.identifier, "code": g.code, "title": g.title}
        for g in rows
    ]

@router.get("/lookup/teachers")
def lookup_teachers(db: Session = Depends(get_db), q: str | None = Query(None), subdivision_id: int | None = Query(None)):
    stmt = select(Teacher)
    if q:
        stmt = stmt.where(or_(Teacher.full_name.ilike(f"%{q}%"), Teacher.email.ilike(f"%{q}%")))
    if subdivision_id:
        stmt = stmt.where(Teacher.subdivision_id == subdivision_id)
    rows = db.scalars(stmt).all()
    return [
        {"id": t.id, "full_name": t.full_name, "email": t.email, "phone": t.phone, "subject": t.subject,
         "subdivision_id": t.subdivision_id}
        for t in rows
    ]

@router.get(
    "/lessons/{lesson_id}/students",
    dependencies=[Depends(require_permission("schedules:read"))],
)
def get_students_for_lesson(
    lesson_id: int,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    lesson = db.get(Lesson, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")

    if not (
        is_admin(me, db)
        or user_has_role(db, me.id, "director")
        or user_has_role(db, me.id, "teacher")
    ):
        raise HTTPException(status_code=403, detail="Not allowed")

    students = (
        db.query(StudentModel)
        .join(User, User.id == StudentModel.user_id)
        .filter(StudentModel.group_id == lesson.group_id)
        .all()
    )

    result = []
    for s in students:
        grades = (
            db.query(Grade)
            .filter(Grade.student_id == s.id, Grade.lesson_id == lesson.id)
            .all()
        )

        final_grade = db.scalar(
            select(Grade.value)
            .where(
                Grade.student_id == s.id,
                Grade.subject_id == lesson.subject_id,
                Grade.grade_type.in_(["итог", "final", "exam", "зачет"]),
            )
            .limit(1)
        )

        result.append(
            {
                "id": s.user.id,
                "full_name": s.user.full_name,
                "email": s.user.email,
                "phone": s.user.phone,
                "record_book": s.record_book,
                "course": s.course,
                "insert_year": s.insert_year,
                "student_id": s.id,
                "final_grade": final_grade,
                "grades": [
                    {
                        "id": g.id,
                        "subject_id": g.subject_id,
                        "subject_code": g.subject.code if g.subject else None,
                        "teacher_id": g.teacher_id,
                        "lesson_id": g.lesson_id,
                        "grade_type": g.grade_type,
                        "value": g.value,
                        "graded_at": g.graded_at,
                        "comment": g.comment,
                    }
                    for g in grades
                ],
            }
        )

    return {
        "lesson_id": lesson.id,
        "group_id": lesson.group_id,
        "group_code": lesson.group.code if lesson.group else None,
        "subject_id": lesson.subject_id,
        "subject_code": lesson.subject.code if lesson.subject else None,
        "subject_title": lesson.subject.title if lesson.subject else None,
        "grade_type": lesson.subject.grade_type if lesson.subject else None,
        "students": result,
    }


@router.get("/lookup/rooms")
def lookup_rooms(db: Session = Depends(get_db), q: str | None = Query(None)):
    stmt = select(Room)
    if q:
        stmt = stmt.where(or_(Room.code.ilike(f"%{q}%"), Room.title.ilike(f"%{q}%")))
    rows = db.scalars(stmt).all()
    return [{"id": r.id, "code": r.code, "title": r.title, "capacity": r.capacity} for r in rows]

@router.get("/lookup/subjects")
def lookup_subjects(db: Session = Depends(get_db), q: str | None = Query(None)):
    stmt = select(Subject).options(
        selectinload(Subject.teachers), selectinload(Subject.primary_teacher)
    )
    if q:
        stmt = stmt.where(or_(Subject.title.ilike(f"%{q}%"), Subject.code.ilike(f"%{q}%")))
    rows = db.scalars(stmt).all()
    return [
        {
            "id": s.id,
            "title": s.title,
            "code": s.code,
            "subject_code": s.code,
            "grade_type": s.grade_type,
            **subject_teacher_payload(s),
        }
        for s in rows
    ]

@router.get("/groups/{group_identifier}")
def get_group(group_identifier: str, db: Session = Depends(get_db)):
    group = None
    if group_identifier.isdigit():
        group = db.get(Group, int(group_identifier))
    if not group:
        group = db.scalar(select(Group).where(Group.code == group_identifier))
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    students = (
        db.query(StudentModel)
        .join(User, User.id == StudentModel.user_id)
        .filter(StudentModel.group_id == group.id)
        .all()
    )

    return {
        "id": group.id,
        "code": group.code,
        "title": group.title,
        "students": [
            {
                "id": s.user.id,
                "full_name": s.user.full_name,
                "email": s.user.email,
                "phone": s.user.phone,
                "record_book": s.record_book,
                "course": s.course,
                "insert_year": s.insert_year,
                "student_id": s.id,
            }
            for s in students
        ],
    }
