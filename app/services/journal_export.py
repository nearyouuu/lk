from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.grade import Student
from app.models.journal import JournalEntry, JournalLesson
from app.models.schedule import Group, Subject
from app.services.journal_service import student_full_name


JOURNAL_EXPORT_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

ATTENDANCE_LABELS = {
    "present": "Б",
    "absent": "Н",
    "late": "О",
    "excused": "У",
}

ATTENDANCE_FILLS = {
    "present": "E2F0D9",
    "absent": "FCE4D6",
    "late": "FFF2CC",
    "excused": "DDEBF7",
}

LESSON_TYPE_LABELS = {
    "lecture": "Лекция",
    "practice": "Практическое занятие",
    "lab": "Лабораторная работа",
}


def _grade_fill(value: str | None) -> str | None:
    return {
        "5": "C6E0B4",
        "4": "E2F0D9",
        "3": "FFF2CC",
        "2": "F4B084",
    }.get((value or "").strip())


def _average_numeric_grade(grades: list[str]) -> float | None:
    values = [int(value) for value in grades if value in {"2", "3", "4", "5"}]
    return round(sum(values) / len(values), 2) if values else None


def build_journal_workbook(
    *,
    group: Group,
    subject: Subject,
    date_from: date,
    date_to: date,
    students: list[Student],
    lessons: list[JournalLesson],
    entries: list[JournalEntry],
) -> bytes:
    """Build a print-ready journal workbook without schedule-specific fields."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Учебный журнал"
    sheet.sheet_view.showGridLines = False

    active_lessons = [lesson for lesson in lessons if lesson.status != "cancelled"]
    entry_map = {(entry.student_id, entry.lesson_id): entry for entry in entries}
    lesson_column_count = len(active_lessons) * 2
    last_column_number = 3 + lesson_column_count + 2
    last_column = get_column_letter(last_column_number)

    dark_blue = "1F4E78"
    blue = "4472C4"
    pale_blue = "D9EAF7"
    thin = Side(style="thin", color="B4C6E7")
    medium = Side(style="medium", color=dark_blue)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells(f"A1:{last_column}1")
    title_cell = sheet["A1"]
    title_cell.value = "Учебный журнал"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=dark_blue)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(f"A2:{last_column}2")
    subtitle = sheet["A2"]
    subject_label = " · ".join(filter(None, [subject.code, subject.title]))
    subtitle.value = f"Группа: {group.code} · Дисциплина: {subject_label}"
    subtitle.font = Font(name="Arial", size=11, bold=True, color=dark_blue)
    subtitle.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 24

    sheet.merge_cells(f"A3:{last_column}3")
    period_cell = sheet["A3"]
    period_cell.value = (
        f"Период: {date_from.strftime('%d.%m.%Y')}–{date_to.strftime('%d.%m.%Y')}"
        f" · Всего часов: {sum(lesson.hours for lesson in active_lessons)}"
    )
    period_cell.font = Font(name="Arial", size=10, italic=True, color="44546A")
    period_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 5
    subheader_row = 6
    fixed_headers = (("№", 6), ("ФИО студента", 30), ("Зачётная книжка", 18))
    for column_number, (label, width) in enumerate(fixed_headers, start=1):
        sheet.merge_cells(
            start_row=header_row,
            start_column=column_number,
            end_row=subheader_row,
            end_column=column_number,
        )
        cell = sheet.cell(header_row, column_number, label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=medium)
        sheet.column_dimensions[get_column_letter(column_number)].width = width

    current_column = 4
    for lesson in active_lessons:
        sheet.merge_cells(
            start_row=header_row,
            start_column=current_column,
            end_row=header_row,
            end_column=current_column + 1,
        )
        date_cell = sheet.cell(
            header_row,
            current_column,
            f"{lesson.lesson_date.strftime('%d.%m.%Y')} · {lesson.hours} ч.",
        )
        date_cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        date_cell.fill = PatternFill("solid", fgColor=blue)
        date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        date_cell.border = Border(bottom=medium)
        for offset, label in enumerate(("Посещ.", "Оценка")):
            cell = sheet.cell(subheader_row, current_column + offset, label)
            cell.font = Font(name="Arial", size=8, bold=True, color=dark_blue)
            cell.fill = PatternFill("solid", fgColor=pale_blue)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
            sheet.column_dimensions[get_column_letter(current_column + offset)].width = 10
        current_column += 2

    summary_headers = (("Пропущено, ч.", 15), ("Средняя оценка", 16))
    for label, width in summary_headers:
        sheet.merge_cells(
            start_row=header_row,
            start_column=current_column,
            end_row=subheader_row,
            end_column=current_column,
        )
        cell = sheet.cell(header_row, current_column, label)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark_blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=medium)
        sheet.column_dimensions[get_column_letter(current_column)].width = width
        current_column += 1

    first_data_row = 7
    sorted_students = sorted(students, key=student_full_name)
    for row_number, student in enumerate(sorted_students, start=first_data_row):
        values = (row_number - first_data_row + 1, student_full_name(student), student.record_book or "")
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column_number, value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                horizontal="left" if column_number == 2 else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = cell_border

        missed_hours = 0
        grades: list[str] = []
        current_column = 4
        for lesson in active_lessons:
            entry = entry_map.get((student.id, lesson.id))
            attendance = entry.attendance if entry else "present"
            grade = (entry.grade or "") if entry else ""
            attendance_cell = sheet.cell(
                row_number, current_column, ATTENDANCE_LABELS.get(attendance, attendance)
            )
            attendance_cell.fill = PatternFill(
                "solid", fgColor=ATTENDANCE_FILLS.get(attendance, "FFFFFF")
            )
            attendance_cell.font = Font(name="Arial", size=9, bold=True)
            attendance_cell.alignment = Alignment(horizontal="center", vertical="center")
            attendance_cell.border = cell_border

            grade_cell = sheet.cell(row_number, current_column + 1, grade)
            grade_cell.font = Font(name="Arial", size=9, bold=bool(grade))
            grade_cell.alignment = Alignment(horizontal="center", vertical="center")
            grade_cell.border = cell_border
            grade_color = _grade_fill(grade)
            if grade_color:
                grade_cell.fill = PatternFill("solid", fgColor=grade_color)
            if grade:
                grades.append(grade)
            if attendance == "absent":
                missed_hours += lesson.hours
            current_column += 2

        missed_cell = sheet.cell(row_number, current_column, missed_hours)
        average_cell = sheet.cell(row_number, current_column + 1, _average_numeric_grade(grades))
        for cell in (missed_cell, average_cell):
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border
        average_cell.number_format = "0.00"
        sheet.row_dimensions[row_number].height = 28

    if not sorted_students:
        sheet.merge_cells(
            start_row=first_data_row,
            start_column=1,
            end_row=first_data_row,
            end_column=last_column_number,
        )
        empty_cell = sheet.cell(first_data_row, 1, "В выбранной группе нет студентов")
        empty_cell.font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = max(first_data_row, sheet.max_row)
    sheet.freeze_panes = "D7"
    sheet.print_title_rows = "1:6"
    sheet.print_title_cols = "A:C"
    sheet.print_area = f"A1:{last_column}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "Страница &P из &N"
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4

    topics_sheet = workbook.create_sheet("Даты и темы")
    topics_sheet.sheet_view.showGridLines = False
    topic_columns = (
        ("№", 6),
        ("Дата", 14),
        ("Часы", 10),
        ("Тип", 24),
        ("Тема", 55),
        ("Комментарий", 40),
    )
    topics_sheet.merge_cells("A1:F1")
    topics_title = topics_sheet["A1"]
    topics_title.value = f"Даты и темы · {group.code} · {subject_label}"
    topics_title.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    topics_title.fill = PatternFill("solid", fgColor=dark_blue)
    topics_title.alignment = Alignment(horizontal="center", vertical="center")
    for column_number, (label, width) in enumerate(topic_columns, start=1):
        cell = topics_sheet.cell(3, column_number, label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        topics_sheet.column_dimensions[get_column_letter(column_number)].width = width
    for row_number, lesson in enumerate(active_lessons, start=4):
        values = (
            row_number - 3,
            lesson.lesson_date,
            lesson.hours,
            LESSON_TYPE_LABELS.get(lesson.lesson_type, lesson.lesson_type),
            lesson.topic_text or "",
            lesson.comment or "",
        )
        for column_number, value in enumerate(values, start=1):
            cell = topics_sheet.cell(row_number, column_number, value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                horizontal="left" if column_number in {5, 6} else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = cell_border
        topics_sheet.cell(row_number, 2).number_format = "dd.mm.yyyy"
        topics_sheet.row_dimensions[row_number].height = 30
    topics_sheet.freeze_panes = "A4"
    topics_sheet.auto_filter.ref = f"A3:F{max(3, topics_sheet.max_row)}"
    topics_sheet.print_title_rows = "1:3"
    topics_sheet.page_setup.orientation = "landscape"
    topics_sheet.page_setup.fitToWidth = 1
    topics_sheet.sheet_properties.pageSetUpPr.fitToPage = True

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
