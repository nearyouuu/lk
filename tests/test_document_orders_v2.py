import os
from io import BytesIO

os.environ["DATABASE_URL"] = "sqlite://"

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.core import deps
from app.core.deps import get_db
from app.core.security import create_access_token
from app.db.base import Base
from app.models import achievement, application, audit, document_order, material, news, profile, role, subject_type, testing  # noqa: F401, E501
from app.models import grade, schedule  # noqa: F401
from app.models.grade import Student
from app.models.role import Role, user_roles
from app.models.schedule import Group
from app.models.user import User
from app.routers import document_orders


def _payload(**overrides):
    payload = {
        "full_name": "Подставное Имя",
        "order_location": "ivanovo_medical_college",
        "department": "nursing",
        "social_protection_information": "nursing_9_full_time_first",
        "study_form": "full_time",
        "group_name": "ЧУЖАЯ-ГРУППА",
        "certificate_type": "scholarship_payment",
        "scholarship_payment_period": "custom",
        "custom_scholarship_payment_period": "  с 01.09.2025 по 31.05.2026  ",
        "place_of_requirement": "  Социальная защита  ",
        "copies_count": 1,
    }
    payload.update(overrides)
    return payload


def _setup():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    db = Session(engine)

    student_role = Role(name="student")
    admin_role = Role(name="administrator")
    group_one = Group(code="СД-21", title="Сестринское дело")
    group_two = Group(code="СД-22", title="Сестринское дело 2")
    owner = User(email="owner@test.kz", password_hash="hash", full_name="Иванов Иван Иванович")
    other = User(email="other@test.kz", password_hash="hash", full_name="Петров Пётр Петрович")
    admin = User(email="admin@test.kz", password_hash="hash", full_name="Администратор")
    db.add_all([student_role, admin_role, group_one, group_two, owner, other, admin])
    db.flush()
    owner_student = Student(user_id=owner.id, group_id=group_one.id)
    other_student = Student(user_id=other.id, group_id=group_two.id)
    db.add_all([owner_student, other_student])
    db.execute(user_roles.insert().values(user_id=owner.id, role_id=student_role.id))
    db.execute(user_roles.insert().values(user_id=other.id, role_id=student_role.id))
    db.execute(user_roles.insert().values(user_id=admin.id, role_id=admin_role.id))
    db.commit()

    app = FastAPI()
    app.include_router(document_orders.router)

    def override_db():
        yield db

    app.dependency_overrides[get_db] = override_db
    deps.has_feature = lambda _feature: (True, None)
    client = TestClient(app)

    def headers(user: User):
        return {"Authorization": f"Bearer {create_access_token(str(user.id))}"}

    return engine, db, client, headers, owner, other, admin


def test_document_order_v2_complete_flow_and_authoritative_profile_fields():
    engine, db, client, headers, owner, other, admin = _setup()
    try:
        created = client.post("/document-orders", json=_payload(), headers=headers(owner))
        assert created.status_code == 201, created.text
        body = created.json()
        order_id = body["id"]
        assert body["full_name"] == "Иванов Иван Иванович"
        assert body["group_name"] == "СД-21"
        assert body["place_of_requirement"] == "Социальная защита"
        assert body["custom_scholarship_payment_period"] == "с 01.09.2025 по 31.05.2026"
        assert body["status"] == "new"
        assert "document_type" not in body

        mine = client.get("/document-orders/me", headers=headers(owner))
        assert mine.status_code == 200
        assert [item["id"] for item in mine.json()] == [order_id]

        all_new = client.get("/document-orders?status=new", headers=headers(admin))
        assert all_new.status_code == 200
        assert [item["id"] for item in all_new.json()] == [order_id]

        patched = client.patch(
            f"/document-orders/{order_id}",
            json={"status": "ready", "comment_admin": "  Кабинет 12  "},
            headers=headers(admin),
        )
        assert patched.status_code == 200
        assert patched.json()["status"] == "ready"
        assert patched.json()["comment_admin"] == "Кабинет 12"

        forbidden = client.post(
            f"/document-orders/{order_id}/confirm-receipt",
            headers=headers(other),
        )
        assert forbidden.status_code == 403
        assert forbidden.json() == {"detail": "Нельзя подтвердить получение чужой заявки"}

        confirmed = client.post(
            f"/document-orders/{order_id}/confirm-receipt",
            headers=headers(owner),
        )
        assert confirmed.status_code == 200
        assert confirmed.json()["status"] == "student_approved"

        owner_delete = client.delete(f"/document-orders/{order_id}", headers=headers(owner))
        assert owner_delete.status_code == 403
        assert owner_delete.json() == {"detail": "Можно удалить только новую заявку"}

        admin_delete = client.delete(f"/document-orders/{order_id}", headers=headers(admin))
        assert admin_delete.status_code == 200
        assert admin_delete.json() == {"status": "deleted", "id": order_id}
    finally:
        db.close()
        engine.dispose()


def test_document_order_v2_returns_string_validation_errors():
    engine, db, client, headers, owner, _other, _admin = _setup()
    try:
        cases = (
            (
                _payload(copies_count=0),
                "Количество экземпляров должно быть от 1 до 10",
            ),
            (
                _payload(order_location="unknown"),
                "Недопустимое значение поля order_location",
            ),
            (
                _payload(
                    certificate_type="education",
                    scholarship_payment_period="3_months",
                    custom_scholarship_payment_period=None,
                ),
                "Для справки об обучении период выплаты должен быть null",
            ),
            (
                _payload(custom_scholarship_payment_period="   "),
                "Для произвольного периода необходимо указать его значение",
            ),
            (
                _payload(
                    scholarship_payment_period="3_months",
                    custom_scholarship_payment_period="лишнее",
                ),
                "Произвольный период должен быть null для выбранного периода выплаты",
            ),
            (
                _payload(full_name="   "),
                "Поле не может быть пустым",
            ),
            (
                _payload(place_of_requirement=123),
                "Поле place_of_requirement должно быть строкой",
            ),
        )
        for payload, expected_detail in cases:
            response = client.post("/document-orders", json=payload, headers=headers(owner))
            assert response.status_code == 422, response.text
            assert response.json() == {"detail": expected_detail}

        missing = _payload()
        missing.pop("place_of_requirement")
        response = client.post("/document-orders", json=missing, headers=headers(owner))
        assert response.status_code == 422
        assert response.json() == {
            "detail": "Обязательное поле не заполнено: place_of_requirement"
        }

        invalid_status = client.get(
            "/document-orders?status=unknown",
            headers=headers(_admin),
        )
        assert invalid_status.status_code == 422
        assert invalid_status.json() == {"detail": "Недопустимое значение поля status"}

        invalid_json = client.post(
            "/document-orders",
            content="{",
            headers={**headers(owner), "Content-Type": "application/json"},
        )
        assert invalid_json.status_code == 422
        assert invalid_json.json() == {"detail": "Некорректный JSON"}
    finally:
        db.close()
        engine.dispose()


def test_book_delivery_filter_export_and_public_link_without_authorization():
    engine, db, client, headers, owner, _other, admin = _setup()
    try:
        created = client.post(
            "/document-orders",
            json={
                "order_type": "book_delivery",
                "request_text": "  Прошу привезти книгу по анатомии  ",
            },
            headers=headers(owner),
        )
        assert created.status_code == 201, created.text
        body = created.json()
        assert body["order_type"] == "book_delivery"
        assert body["request_text"] == "Прошу привезти книгу по анатомии"
        assert body["full_name"] == "Иванов Иван Иванович"
        assert body["group_name"] == "СД-21"
        assert body["certificate_type"] is None
        assert body["copies_count"] is None

        certificate = client.post(
            "/document-orders",
            json=_payload(),
            headers=headers(owner),
        )
        assert certificate.status_code == 201, certificate.text

        filtered = client.get(
            "/document-orders?order_type=book_delivery&status=new&q=анатомии",
            headers=headers(admin),
        )
        assert filtered.status_code == 200, filtered.text
        assert [item["id"] for item in filtered.json()] == [body["id"]]

        forbidden_export = client.get("/document-orders/export", headers=headers(owner))
        assert forbidden_export.status_code == 403

        direct_export = client.get(
            "/document-orders/export?status=new&q=анатомии",
            headers=headers(admin),
        )
        assert direct_export.status_code == 200, direct_export.text
        assert direct_export.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        direct_workbook = load_workbook(BytesIO(direct_export.content), read_only=True)
        direct_rows = list(direct_workbook["Заявки на книги"].iter_rows(values_only=True))
        assert len(direct_rows) == 2
        assert direct_rows[1][4] == "Прошу привезти книгу по анатомии"

        forbidden_link = client.post(
            "/document-orders/export-links",
            json={"status": "new"},
            headers=headers(owner),
        )
        assert forbidden_link.status_code == 403

        link_response = client.post(
            "/document-orders/export-links",
            json={"status": "new", "q": "анатомии", "expires_in_hours": 24},
            headers=headers(admin),
        )
        assert link_response.status_code == 200, link_response.text
        public_url = link_response.json()["url"]
        assert "/document-orders/public-export/" in public_url

        public_export = client.get(public_url)
        assert public_export.status_code == 200, public_export.text
        public_workbook = load_workbook(BytesIO(public_export.content), read_only=True)
        public_rows = list(public_workbook["Заявки на книги"].iter_rows(values_only=True))
        assert direct_rows == public_rows

        invalid_link = client.get(f"{public_url}broken")
        assert invalid_link.status_code == 404
        assert invalid_link.json() == {
            "detail": "Ссылка на выгрузку недействительна или истекла"
        }
    finally:
        db.close()
        engine.dispose()


def test_book_delivery_request_text_validation():
    engine, db, client, headers, owner, _other, _admin = _setup()
    try:
        empty = client.post(
            "/document-orders",
            json={"order_type": "book_delivery", "request_text": "   "},
            headers=headers(owner),
        )
        assert empty.status_code == 422
        assert empty.json() == {"detail": "Текст заявки на книгу не может быть пустым"}

        too_long = client.post(
            "/document-orders",
            json={"order_type": "book_delivery", "request_text": "а" * 2001},
            headers=headers(owner),
        )
        assert too_long.status_code == 422
        assert too_long.json() == {
            "detail": "Текст заявки на книгу не должен превышать 2000 символов"
        }
    finally:
        db.close()
        engine.dispose()
