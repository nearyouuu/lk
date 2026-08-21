import os
from datetime import datetime
from io import BytesIO

os.environ["DATABASE_URL"] = "sqlite://"

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.models import (  # noqa: F401
    achievement,
    application,
    audit,
    document_order,
    grade,
    material,
    news,
    profile,
    schedule,
    subject_type,
    testing,
)
from app.models.schedule import Group, Lesson, Room, Subject, Teacher
from app.routers.schedules import SCHEDULE_EXPORT_MEDIA_TYPE, export_schedule


def _engine():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    return engine


def _export(db: Session, **filters):
    params = {
        "group_code": None,
        "date_from": None,
        "date_to": None,
        "teacher_full_name": None,
        "subject_code": None,
        "subject_title": None,
        "room_code": None,
        "subject_type": None,
    }
    params.update(filters)
    return export_schedule(db=db, **params)


def test_schedule_export_is_print_ready_xlsx():
    engine = _engine()
    with Session(engine) as db:
        group = Group(code="ИС-101", title="ИС-101")
        subject = Subject(code="MATH-01", title="Математика")
        teacher = Teacher(full_name="Иванов Иван Иванович")
        room = Room(code="204", title="204")
        db.add_all([group, subject, teacher, room])
        db.flush()
        db.add(
            Lesson(
                group_id=group.id,
                subject_id=subject.id,
                teacher_id=teacher.id,
                room_id=room.id,
                lesson_number=1,
                starts_at=datetime(2026, 8, 24, 9, 0),
                ends_at=datetime(2026, 8, 24, 10, 30),
                subject_type="lecture",
                topic="Квадратные уравнения",
                notes="Взять тетрадь",
            )
        )
        db.commit()

        response = _export(db, group_code="ИС-101")

    assert response.media_type == SCHEDULE_EXPORT_MEDIA_TYPE
    assert "attachment" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.body))
    sheet = workbook["Расписание"]
    assert sheet["A1"].value == "Расписание занятий"
    assert sheet["A2"].value == "Группа: ИС-101"
    assert sheet["A5"].value == datetime(2026, 8, 24)
    assert sheet["B5"].value == "Понедельник"
    assert sheet["F5"].value == "Математика"
    assert sheet["G5"].value == "Лекция"
    assert sheet["H5"].value == "Иванов Иван Иванович"
    assert sheet["J5"].value == "Квадратные уравнения"
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.print_title_rows == "$1:$4"
    assert sheet.freeze_panes == "A5"
    engine.dispose()


def test_schedule_export_handles_empty_result():
    engine = _engine()
    with Session(engine) as db:
        response = _export(db, group_code="НЕ-СУЩЕСТВУЕТ")

    workbook = load_workbook(BytesIO(response.body))
    assert workbook["Расписание"]["A5"].value == "По выбранным фильтрам занятий нет"
    engine.dispose()
