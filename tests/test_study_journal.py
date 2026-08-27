import os
from datetime import date, datetime, time
from io import BytesIO

os.environ["DATABASE_URL"] = "sqlite://"

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.models import (  # noqa: F401
    achievement,
    application,
    audit,
    document_order,
    grade,
    journal,
    material,
    news,
    profile,
    role,
    schedule,
    subject_type,
    testing,
    user,
)
from app.models.grade import Student
from app.models.journal import JournalEntry, JournalPeriod
from app.models.role import Permission, Role, role_permissions, user_roles
from app.models.schedule import Group, Lesson, Subject, Teacher
from app.models.user import User
from app.routers.admin_journal import (
    create_journal_assignment,
    create_subject_topic,
    lock_journal_period,
)
from app.routers.journal import (
    create_journal_lesson,
    export_journal_excel,
    get_journal,
    journal_catalog,
    journal_group_students,
    put_journal_entries_batch,
    put_journal_entry,
)
from app.routers.control_points import (
    control_point_statement,
    generate_control_points,
    put_control_point_score,
)
from app.schemas.journal import (
    JournalAssignmentCreate,
    JournalBatchPut,
    JournalEntryPut,
    JournalLessonCreate,
    SubjectTopicCreate,
    ControlPointScorePut,
    ControlPointsGenerate,
)
from app.services.journal_service import JournalAPIError


def _engine():
    return create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )


def _fixture(db: Session):
    admin_role = Role(name="administrator")
    teacher_role = Role(name="teacher")
    student_role = Role(name="student")
    admin = User(email="journal-admin@test.kz", password_hash="hash", full_name="Admin")
    teacher_user = User(
        email="journal-teacher@test.kz", password_hash="hash", full_name="Teacher"
    )
    group = Group(code="J-21", title="Journal group")
    subject = Subject(code="JOURNAL-01", title="Journal subject", grade_type="exam")
    db.add_all([admin_role, teacher_role, student_role, admin, teacher_user, group, subject])
    db.flush()
    db.execute(
        user_roles.insert(),
        [
            {"user_id": admin.id, "role_id": admin_role.id},
            {"user_id": teacher_user.id, "role_id": teacher_role.id},
        ],
    )
    for code in ("journal.read", "journal.lesson.write", "journal.entry.write"):
        permission = Permission(code=code, description=code)
        db.add(permission)
        db.flush()
        db.execute(
            role_permissions.insert().values(
                role_id=teacher_role.id, permission_id=permission.id
            )
        )
        if code == "journal.read":
            db.execute(
                role_permissions.insert().values(
                    role_id=student_role.id, permission_id=permission.id
                )
            )
    teacher = Teacher(user_id=teacher_user.id, full_name="Teacher")
    db.add(teacher)
    students = []
    for index in range(2):
        student_user = User(
            email=f"journal-student-{index}@test.kz",
            password_hash="hash",
            full_name=f"Student {index}",
        )
        db.add(student_user)
        db.flush()
        db.execute(
            user_roles.insert().values(
                user_id=student_user.id, role_id=student_role.id
            )
        )
        student = Student(
            user_id=student_user.id,
            group_id=group.id,
            record_book=f"J-{index}",
        )
        db.add(student)
        students.append(student)
    db.commit()
    return admin, teacher_user, teacher, group, subject, students


def test_independent_journal_flow_and_period_lock():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, teacher_user, teacher, group, subject, students = _fixture(db)
        assignment = create_journal_assignment(
            JournalAssignmentCreate(
                teacher_id=teacher.id,
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
            ),
            db,
            admin,
        )
        assert assignment["is_active"] is True

        topic = create_subject_topic(
            subject.id,
            SubjectTopicCreate(title="Independent topic", sort_order=0),
            db,
            admin,
        )
        lesson = create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 8, 24),
                starts_at=time(9, 0),
                ends_at=time(10, 30),
                type="practice",
                topic_id=topic["id"],
                status="published",
            ),
            "journal-idempotency-1",
            "req_test",
            db,
            teacher_user,
        )
        assert lesson["topic_text"] == "Independent topic"
        assert lesson["schedule_lesson_id"] is None

        payload = get_journal(
            group.id,
            subject.id,
            date(2026, 8, 1),
            date(2026, 12, 31),
            db,
            teacher_user,
        )
        assert len(payload["students"]) == 2
        assert len(payload["lessons"]) == 1

        saved = put_journal_entry(
            lesson["id"],
            students[0].id,
            JournalEntryPut(attendance="late", grade="4", comment="10 min", version=0),
            "req_entry",
            db,
            teacher_user,
        )
        assert saved["version"] == 1

        with pytest.raises(JournalAPIError) as conflict:
            put_journal_entry(
                lesson["id"],
                students[0].id,
                JournalEntryPut(attendance="present", grade="5", version=0),
                None,
                db,
                teacher_user,
            )
        assert conflict.value.status_code == 409

        period = db.get(JournalPeriod, payload["period"]["id"])
        lock_journal_period(period.id, db, admin)
        with pytest.raises(JournalAPIError) as locked:
            put_journal_entry(
                lesson["id"],
                students[0].id,
                JournalEntryPut(attendance="present", grade="5", version=1),
                None,
                db,
                teacher_user,
            )
        assert locked.value.status_code == 423
    engine.dispose()


def test_batch_save_has_partial_success():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, teacher_user, teacher, group, subject, students = _fixture(db)
        create_journal_assignment(
            JournalAssignmentCreate(
                teacher_id=teacher.id,
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
            ),
            db,
            admin,
        )
        lesson = create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 9, 1),
                type="lecture",
                topic_text="Manual topic",
            ),
            None,
            None,
            db,
            teacher_user,
        )
        put_journal_entry(
            lesson["id"],
            students[0].id,
            JournalEntryPut(attendance="present", grade="5", version=0),
            None,
            db,
            teacher_user,
        )

        result = put_journal_entries_batch(
            lesson["id"],
            JournalBatchPut(
                entries=[
                    {
                        "student_id": students[0].id,
                        "attendance": "late",
                        "grade": "4",
                        "version": 0,
                    },
                    {
                        "student_id": students[1].id,
                        "attendance": "absent",
                        "grade": None,
                        "version": 0,
                    },
                ]
            ),
            None,
            db,
            teacher_user,
        )
        assert [row["student_id"] for row in result["updated"]] == [students[1].id]
        assert result["failed"][0]["code"] == "JOURNAL_VERSION_CONFLICT"
        assert db.query(JournalEntry).filter_by(lesson_id=lesson["id"]).count() == 2
    engine.dispose()


def test_control_points_formula_attendance_and_project_limit():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, teacher_user, teacher, group, subject, students = _fixture(db)
        create_journal_assignment(
            JournalAssignmentCreate(
                teacher_id=teacher.id,
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
            ),
            db,
            admin,
        )
        lesson_ids = []
        for index in range(17):
            lesson = create_journal_lesson(
                JournalLessonCreate(
                    group_id=group.id,
                    subject_id=subject.id,
                    date=date(2026, 8, 1 + index),
                    starts_at=time(9, 0),
                    ends_at=time(13, 0),
                    hours=4,
                    type="practice",
                    topic_text=f"Topic {index + 1}",
                ),
                f"kt-lesson-{index}",
                None,
                db,
                teacher_user,
            )
            lesson_ids.append(lesson["id"])

        generated = generate_control_points(
            ControlPointsGenerate(
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
                total_practical_hours=68,
            ),
            None,
            db,
            teacher_user,
        )
        assert generated["lesson_count"] == 17
        assert generated["interval"] == 23
        assert [row["planned_lesson_number"] for row in generated["items"]] == [6, 12, 17]
        assert [row["planned_hours"] for row in generated["items"]] == [23, 46, 68]
        assert [row["base_max"] for row in generated["items"]] == [23.0, 23.0, 24.0]
        assert all("teacher_id" not in row for row in generated["items"])

        for index, lesson_id in enumerate(lesson_ids[:6]):
            put_journal_entry(
                lesson_id,
                students[0].id,
                JournalEntryPut(
                    attendance="absent" if index == 0 else "present",
                    version=0,
                ),
                None,
                db,
                teacher_user,
            )

        first_point = generated["items"][0]
        first_score = put_control_point_score(
            first_point["id"],
            students[0].id,
            ControlPointScorePut(
                current_score=18,
                project_score=15,
                attendance_score=None,
                version=1,
            ),
            None,
            db,
            teacher_user,
        )
        assert first_score["eligible_hours"] == 24
        assert first_score["attended_hours"] == 20
        assert first_score["attendance_score"] == 2.5
        assert first_score["total_score"] == 35.5

        second_point = generated["items"][1]
        with pytest.raises(JournalAPIError) as project_limit:
            put_control_point_score(
                second_point["id"],
                students[0].id,
                ControlPointScorePut(
                    current_score=20,
                    project_score=6,
                    attendance_score=None,
                    version=1,
                ),
                None,
                db,
                teacher_user,
            )
        assert project_limit.value.detail["error"]["code"] == "JOURNAL_PROJECT_SCORE_LIMIT"

        statement = control_point_statement(
            group.id,
            subject.id,
            2026,
            "autumn",
            db,
            teacher_user,
        )
        assert statement["maximums"] == {
            "current": 60,
            "attendance": 10,
            "project": 20,
            "semester_total": 90,
            "control_points": [23, 23, 24],
        }
        student_row = next(
            row for row in statement["items"] if row["student"]["id"] == students[0].id
        )
        assert student_row["semester_total"] == 35.5
    engine.dispose()


def test_control_points_are_rejected_for_industrial_practice():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, teacher_user, teacher, group, subject, _students = _fixture(db)
        create_journal_assignment(
            JournalAssignmentCreate(
                teacher_id=teacher.id,
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
            ),
            db,
            admin,
        )
        with pytest.raises(JournalAPIError) as exc_info:
            generate_control_points(
                ControlPointsGenerate(
                    group_id=group.id,
                    subject_id=subject.id,
                    academic_year=2026,
                    semester="autumn",
                    total_practical_hours=68,
                    study_component="industrial_practice",
                ),
                None,
                db,
                teacher_user,
            )
        assert exc_info.value.detail["error"]["code"] == "JOURNAL_CONTROL_POINTS_NOT_APPLICABLE"
    engine.dispose()


def test_assigned_teachers_share_the_same_journal_lessons():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, first_user, first_teacher, group, subject, students = _fixture(db)
        teacher_role = db.scalar(select(Role).where(Role.name == "teacher"))
        second_user = User(
            email="journal-second-teacher@test.kz",
            password_hash="hash",
            full_name="Second Teacher",
        )
        db.add(second_user)
        db.flush()
        db.execute(
            user_roles.insert().values(user_id=second_user.id, role_id=teacher_role.id)
        )
        second_teacher = Teacher(user_id=second_user.id, full_name="Second Teacher")
        db.add(second_teacher)
        db.commit()

        for teacher in (first_teacher, second_teacher):
            create_journal_assignment(
                JournalAssignmentCreate(
                    teacher_id=teacher.id,
                    group_id=group.id,
                    subject_id=subject.id,
                    academic_year=2026,
                    semester="autumn",
                ),
                db,
                admin,
            )

        lesson = create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 9, 10),
                type="practice",
                topic_text="Shared lesson",
            ),
            None,
            None,
            db,
            first_user,
        )

        visible = get_journal(
            group.id,
            subject.id,
            date(2026, 9, 1),
            date(2026, 9, 30),
            db,
            second_user,
        )
        assert [row["id"] for row in visible["lessons"]] == [lesson["id"]]

        saved = put_journal_entry(
            lesson["id"],
            students[0].id,
            JournalEntryPut(attendance="present", grade="3", version=0),
            None,
            db,
            second_user,
        )
        assert saved["grade"] == "3"

        admin_update = put_journal_entry(
            lesson["id"],
            students[0].id,
            JournalEntryPut(attendance="present", grade="4", version=saved["version"]),
            None,
            db,
            admin,
        )
        assert admin_update["grade"] == "4"
    engine.dispose()


def test_schedule_does_not_grant_journal_access():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        _admin, teacher_user, teacher, group, subject, _students = _fixture(db)
        db.add(
            Lesson(
                group_id=group.id,
                subject_id=subject.id,
                teacher_id=teacher.id,
                lesson_number=1,
                starts_at=datetime(2026, 9, 1, 9, 0),
                ends_at=datetime(2026, 9, 1, 10, 30),
                subject_type="practice",
            )
        )
        db.commit()

        with pytest.raises(JournalAPIError) as denied:
            get_journal(
                group.id,
                subject.id,
                date(2026, 9, 1),
                date(2026, 9, 30),
                db,
                teacher_user,
            )
        assert denied.value.detail["error"]["code"] == "JOURNAL_ACCESS_DENIED"
    engine.dispose()


def test_journal_excel_export_contains_hours_attendance_and_grades():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, teacher_user, teacher, group, subject, students = _fixture(db)
        create_journal_assignment(
            JournalAssignmentCreate(
                teacher_id=teacher.id,
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
            ),
            db,
            admin,
        )
        first = create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 9, 1),
                hours=4,
                topic_text="Практическая работа № 1",
            ),
            None,
            None,
            db,
            teacher_user,
        )
        second = create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 9, 8),
                hours=2,
            ),
            None,
            None,
            db,
            teacher_user,
        )
        put_journal_entry(
            first["id"],
            students[0].id,
            JournalEntryPut(attendance="absent", grade="4", version=0),
            None,
            db,
            teacher_user,
        )
        put_journal_entry(
            second["id"],
            students[0].id,
            JournalEntryPut(attendance="late", grade="5", version=0),
            None,
            db,
            teacher_user,
        )

        response = export_journal_excel(
            group.id,
            subject.id,
            date(2026, 9, 1),
            date(2026, 12, 31),
            db,
            teacher_user,
        )

        assert response.media_type.startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "filename*=UTF-8''" in response.headers["content-disposition"]
        workbook = load_workbook(BytesIO(response.body), data_only=False)
        assert workbook.sheetnames == ["Учебный журнал", "Даты и темы"]
        sheet = workbook["Учебный журнал"]
        assert sheet["D5"].value == "01.09.2026 · 4 ч."
        assert sheet["D6"].value == "Посещ."
        assert sheet["E6"].value == "Оценка"
        assert sheet["D7"].value == "Н"
        assert sheet["E7"].value == "4"
        assert sheet["F7"].value == "О"
        assert sheet["G7"].value == "5"
        assert sheet["H7"].value == 4
        assert sheet["I7"].value == 4.5
        assert workbook["Даты и темы"]["D4"].value == "Практическая работа № 1"
    engine.dispose()


def test_student_sees_only_own_read_only_journal_by_subject():
    engine = _engine()
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        admin, teacher_user, teacher, group, subject, students = _fixture(db)
        create_journal_assignment(
            JournalAssignmentCreate(
                teacher_id=teacher.id,
                group_id=group.id,
                subject_id=subject.id,
                academic_year=2026,
                semester="autumn",
            ),
            db,
            admin,
        )
        published = create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 9, 1),
                hours=4,
                topic_text="Опубликованная тема",
                status="published",
            ),
            None,
            None,
            db,
            teacher_user,
        )
        create_journal_lesson(
            JournalLessonCreate(
                group_id=group.id,
                subject_id=subject.id,
                date=date(2026, 9, 8),
                hours=2,
                topic_text="Черновик",
                status="draft",
            ),
            None,
            None,
            db,
            teacher_user,
        )
        for student, grade_value in zip(students, ("5", "2"), strict=True):
            put_journal_entry(
                published["id"],
                student.id,
                JournalEntryPut(attendance="present", grade=grade_value, version=0),
                None,
                db,
                teacher_user,
            )

        student_user = students[0].user
        catalog = journal_catalog(2026, "autumn", None, db, student_user)
        assert catalog["access_scope"] == "self"
        assert [row["id"] for row in catalog["groups"]] == [group.id]
        assert [row["id"] for row in catalog["groups"][0]["subjects"]] == [subject.id]

        group_students = journal_group_students(
            group.id, date(2026, 9, 1), db, student_user
        )
        assert [row["id"] for row in group_students["items"]] == [students[0].id]

        payload = get_journal(
            group.id,
            subject.id,
            date(2026, 9, 1),
            date(2026, 12, 31),
            db,
            student_user,
        )
        assert payload["access_scope"] == "self"
        assert payload["permissions"] == {
            "can_edit": False,
            "can_manage_topics": False,
        }
        assert [row["id"] for row in payload["students"]] == [students[0].id]
        assert [row["id"] for row in payload["lessons"]] == [published["id"]]
        assert [row["student_id"] for row in payload["entries"]] == [students[0].id]
        assert payload["entries"][0]["grade"] == "5"

        export_response = export_journal_excel(
            group.id,
            subject.id,
            date(2026, 9, 1),
            date(2026, 12, 31),
            db,
            student_user,
        )
        exported_book = load_workbook(BytesIO(export_response.body), read_only=True)
        exported_sheet = exported_book["Учебный журнал"]
        assert exported_sheet["B7"].value == "Student 0"
        assert exported_sheet["B8"].value is None
        assert exported_sheet["E7"].value == "5"

        with pytest.raises(JournalAPIError) as write_denied:
            put_journal_entry(
                published["id"],
                students[0].id,
                JournalEntryPut(attendance="absent", grade="3", version=1),
                None,
                db,
                student_user,
            )
        assert write_denied.value.status_code == 403

        other_group = Group(code="J-22", title="Other group")
        db.add(other_group)
        db.commit()
        with pytest.raises(JournalAPIError) as other_group_denied:
            get_journal(
                other_group.id,
                subject.id,
                date(2026, 9, 1),
                date(2026, 12, 31),
                db,
                student_user,
            )
        assert other_group_denied.value.status_code == 403
    engine.dispose()
